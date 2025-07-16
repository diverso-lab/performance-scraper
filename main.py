import requests
import lxml
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Obtener integracion continua si/no
def obtainCI(direccion):
    actions_url = f"https://github.com/{direccion}/actions"
    github_actions_page = requests.get(actions_url)
    actions_soup = BeautifulSoup(github_actions_page.content, 'html.parser')

    menu = actions_soup.find_all('span', {"class": 'ActionListItem-label ActionListItem-label--truncate'})

    enabled_CI = False
    for item in menu:
        # TODO: darle una vuelta a la comprobacion esta
        if "ci" in item.text.strip().lower():
            enabled_CI = True
            break
    return enabled_CI


# Obtener numero labels
def obtainLabelsCount(direccion):
    labels_url = f"https://github.com/{direccion}/labels"
    labels_page = requests.get(labels_url)
    labels_soup = BeautifulSoup(labels_page.content, 'html.parser')

    return labels_soup.find('span', {"class": 'js-labels-count'}).text.strip()


# Obtener dependencias (el proyecto A depende de X proyectos)
def obtainDependencies(direccion):
    dependencies_url = f"https://github.com/{direccion}/network/dependencies"
    dependencies_page = requests.get(dependencies_url)
    dependencies_soup = BeautifulSoup(dependencies_page.content, 'html.parser')

    return int(dependencies_soup.find('div', {"class": 'd-flex flex-items-center gap-1 text-bold'}).text.strip().split(" ")[0].replace(",", ""))

# Obtener dependencias (X proyectos dependen del proyecto A)
def obtainDependants(direccion):
    dependencies_url = f"https://github.com/{direccion}/network/dependents"
    dependencies_page = requests.get(dependencies_url)
    dependencies_soup = BeautifulSoup(dependencies_page.content, 'html.parser')

    return int(dependencies_soup.find('a', {"class": 'btn-link selected'}).text.strip().split(" ")[0].replace(",", "").replace("\n", ""))


# Obtener community standards cumplimentados
def obtainCommunityStandards(direccion):
    community_standards_url = f"https://github.com/{direccion}/community"
    community_standards_page = requests.get(community_standards_url)
    community_standards_soup = BeautifulSoup(community_standards_page.content, 'html.parser')

# Nos quedamos con la lista de iconos
    item_list = community_standards_soup.find('ul', {"class": "Box"})
    item_list = item_list.find_all('svg')

# Creamos una lista de booleanos y vamos añadiendo en funcion de cada icono
# El orden siempre va a ser: 
# 1. Description, 2. README, 3. Code of Conduct, 4. Contributing, 5. License, 6. Security Policy,
# 7. Issue Template, 8. Pull Request Template, 9. Repository admins accept content reports
    community_standards = []
    for item in item_list:
        if "Added" in item['aria-label']:
            community_standards.append(True)
        else:
            community_standards.append(False)
    return f"{community_standards.count(True)}/9"

proyectos = {}
def textToData():
    fichero = open("proyectos.txt")
    lineas = fichero.readlines()
    for i in lineas:
        direccion = i.removesuffix("\n")
        print(i + ": " + direccion)
        enabled_CI = obtainCI(direccion)
        labels_count = obtainLabelsCount(direccion)
        #depends_on = obtainDependants(direccion)
        #depends_of = obtainDependencies(direccion)
        community_standards = obtainCommunityStandards(direccion)
        proyectos.update({direccion:[enabled_CI,labels_count, community_standards]})#depends_of,depends_on

def dataToXlsx():
    wb = load_workbook("sample.xlsx")
    ws = wb["Hoja1"]
    i = 2
    ws['A1'] = 'Direccion'
    ws['B1'] = 'CI'
    ws['C1'] = 'Labels count'
    #ws['D1'] = 'Depends on'
    #ws['E1'] = 'Depends of'
    ws['F1'] = 'Community Standards'
    for key in proyectos:
        ws[f'A{i}'] = key
        ws[f'F{i}'] = proyectos[key].pop()
        #ws[f'E{i}'] = proyectos[key].pop()
        #ws[f'D{i}'] = proyectos[key].pop()
        ws[f'C{i}'] = proyectos[key].pop()
        ws[f'B{i}'] = proyectos[key].pop()
        i+=1
    wb.save("sample.xlsx")

textToData()
dataToXlsx()
