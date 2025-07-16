import requests
from openpyxl import load_workbook




token = "YOUR-TOKEN"

# Set the headers
headers = {
    "Authorization": f"token {token}",
    "Accept": "application/vnd.github.v3+json"
}


def languagePercentage(direccion):
    response = requests.get(f"  https://api.github.com/repos/{direccion}/languages", headers=headers)
    languagesDic = response.json()
    languageDicValues = languagesDic.values()
    sortedDic = sorted(languageDicValues, key = lambda language: language, reverse=True)
    total = sum(languageDicValues)
    percentageList = []
    language = "Several"
    for value in sortedDic:
        percentage = value*1.0/total*1.0 * 100
        percentageList.append(percentage)
        if percentage >= 75:
            language = list(languagesDic.keys())[list(languagesDic.values()).index(value)]
    return language

def getWorkflows(direccion):
    response2 = requests.get(f"  https://api.github.com/repos/{direccion}/actions/workflows", headers=headers)
    totalWorkflows = response2.json()["total_count"]
    return totalWorkflows

def getStarCount(direccion):
    response3 = requests.get(f"  https://api.github.com/repos/{direccion}", headers=headers)
    starsCount = response3.json()["stargazers_count"]
    return starsCount

def getForkCount(direccion):
    response3 = requests.get(f"  https://api.github.com/repos/{direccion}", headers=headers)
    forksCount = response3.json()["forks_count"]
    return forksCount

def getWachersCount(direccion):
    response3 = requests.get(f"  https://api.github.com/repos/{direccion}", headers=headers)
    watchersCount = response3.json()["subscribers_count"]
    return watchersCount


def getContributors(direccion):
    response4 = requests.get(f" https://api.github.com/repos/{direccion}/contributors?per_page=1&anon=true", headers=headers)
    parts = response4.headers.get("Link").split(",")
    return parts[1].split("true&page=")[1].split(">")[0]

proyectos = {}
def textToData():
    fichero = open("proyectos.txt")
    lineas = fichero.readlines()
    for i in lineas:
        direccion = i.removesuffix("\n")
        print(i + ": " + direccion)
        language = languagePercentage(direccion)
        stars = getStarCount(direccion)
        watchers = getWachersCount(direccion)
        contributors = getContributors(direccion)
        forks = getForkCount(direccion)
        workflows = getWorkflows(direccion)
        proyectos.update({direccion:[language,stars,watchers,contributors,forks,workflows]})

def dataToXlsx():
    wb = load_workbook("sample.xlsx")
    ws = wb["Hoja1"]
    i = 2
    ws['G1'] = 'Lenguaje'
    ws['H1'] = 'Estrellas'
    ws['I1'] = 'Watchers'
    ws['J1'] = 'Contribuidores'
    ws['K1'] = 'Forks'
    ws['L1'] = 'Workflows'
    for key in proyectos:
        ws[f'L{i}'] = proyectos[key].pop()
        ws[f'K{i}'] = proyectos[key].pop()
        ws[f'J{i}'] = proyectos[key].pop()
        ws[f'I{i}'] = proyectos[key].pop()
        ws[f'H{i}'] = proyectos[key].pop()
        ws[f'G{i}'] = proyectos[key].pop()
        i+=1
    wb.save("sample.xlsx")

textToData()
dataToXlsx()
